"""
pipeline_server.py
==================
Servidor FastAPI que executa o pipeline completo de análise de crédito.
Roda no Railway (ou qualquer servidor Python) — NÃO no Supabase.

Recebe { analysis_id } via POST /run-pipeline
Busca arquivos no Supabase Storage, roda E1→E2→E3→E4→montador, salva resultados.

Dependências: ver requirements.txt
Módulos locais: calculadora.py, montador.py, api_client.py (mesmos arquivos, sem alteração)
"""

from __future__ import annotations

import asyncio
import base64
import json
import os
import re
import shutil
import subprocess
import tempfile
import time
from pathlib import Path
from typing import Any

import httpx
from fastapi import BackgroundTasks, FastAPI, Header, HTTPException, Request
from fastapi.responses import JSONResponse
from pydantic import BaseModel

# ── Módulos locais (mesmos arquivos Python — sem alteração) ─────────────────
from calculadora import calcular
from montador import montar_payload, validar_payload, validar_p2
from api_client import enviar_payload, inspecionar_payload

# ── Configuração via variáveis de ambiente ──────────────────────────────────
ANTHROPIC_API_KEY     = os.environ["ANTHROPIC_API_KEY"]
SUPABASE_URL          = os.environ["SUPABASE_URL"]
SUPABASE_SERVICE_KEY  = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
LOVABLE_API_KEY       = os.environ.get("LOVABLE_API_KEY", "")
PIPELINE_SECRET       = os.environ.get("PIPELINE_SECRET", "")  # segurança básica entre Edge Fn e servidor

STORAGE_BUCKET        = "analysis-files"
API_URL               = "https://api.anthropic.com/v1/messages"

MODELO_E1             = "claude-haiku-4-5-20251001"
MODELO_E2             = "claude-haiku-4-5-20251001"
MODELO_E4             = "claude-sonnet-4-20250514"

MAX_TOKENS_E1         = 8_000
MAX_TOKENS_E2         = 16_000
MAX_TOKENS_E4         = 16_000
CHUNK_DOCS            = 1         # 1 doc por vez — seguro para PDFs grandes
SLEEP_ENTRE_LOTES     = 30        # segundos entre lotes
TIMEOUT               = 300       # segundos por chamada à API Anthropic
LIMITE_KB_PDF         = 800       # PDFs maiores que isso são comprimidos

app = FastAPI(title="Pipeline de Análise de Crédito")


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS SUPABASE
# ══════════════════════════════════════════════════════════════════════════════

def _sb_headers() -> dict:
    return {
        "apikey":        SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type":  "application/json",
        "Prefer":        "return=representation",
    }


async def _sb_get(table: str, filters: str) -> list[dict]:
    """GET numa tabela Supabase via REST."""
    url = f"{SUPABASE_URL}/rest/v1/{table}?{filters}"
    async with httpx.AsyncClient(timeout=30) as client:
        r = await client.get(url, headers=_sb_headers())
        r.raise_for_status()
        return r.json()


async def _sb_patch(table: str, id_: str, data: dict) -> None:
    """PATCH numa linha Supabase via REST."""
    url = f"{SUPABASE_URL}/rest/v1/{table}?id=eq.{id_}"
    async with httpx.AsyncClient(timeout=30) as client:
        r = await client.patch(url, headers=_sb_headers(), json=data)
        r.raise_for_status()


async def _sb_download(storage_path: str) -> bytes:
    """Baixa um arquivo do Supabase Storage."""
    url = f"{SUPABASE_URL}/storage/v1/object/{STORAGE_BUCKET}/{storage_path}"
    async with httpx.AsyncClient(timeout=120) as client:
        r = await client.get(url, headers=_sb_headers())
        r.raise_for_status()
        return r.content


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS ANTHROPIC  (portados do pipeline original — lógica idêntica)
# ══════════════════════════════════════════════════════════════════════════════

def _anthropic_headers(cache: bool = False) -> dict:
    h = {
        "Content-Type":      "application/json",
        "x-api-key":         ANTHROPIC_API_KEY,
        "anthropic-version": "2023-06-01",
    }
    if cache:
        h["anthropic-beta"] = "prompt-caching-2024-07-31"
    return h


def _juntar_texto(content: list) -> str:
    return "\n".join(b["text"] for b in content if b.get("type") == "text")


def _extrair_json(texto: str) -> dict:
    """Extrai o primeiro JSON válido — lógica idêntica ao pipeline original."""
    for tentativa in [
        lambda t: json.loads(t.strip()),
        lambda t: json.loads(re.sub(r"```(?:json)?\s*|```\s*$", "", t, flags=re.M).strip()),
    ]:
        try:
            return tentativa(texto)
        except json.JSONDecodeError:
            pass
    ini, prof = texto.find("{"), 0
    if ini == -1:
        raise ValueError("Nenhum JSON encontrado na resposta")
    for i, c in enumerate(texto[ini:], ini):
        prof += (c == "{") - (c == "}")
        if prof == 0:
            return json.loads(texto[ini : i + 1])
    raise ValueError(f"JSON incompleto — resposta truncada em {len(texto)} chars\nÚltimos 300: {texto[-300:]}")


async def _chamar_api(
    mensagens: list[dict],
    max_tokens: int,
    tools: list[dict] | None = None,
    cache: bool = False,
    modelo: str = MODELO_E4,
) -> tuple[str, dict]:
    """
    Chama a API Anthropic — lógica idêntica ao pipeline original.
    Suporta tool_use (web_search) com loop automático e retry em 429.
    """
    body: dict = {"model": modelo, "max_tokens": max_tokens, "messages": mensagens}
    if tools:
        body["tools"] = tools

    uso_total = {
        "input_tokens": 0, "output_tokens": 0,
        "cache_creation_input_tokens": 0, "cache_read_input_tokens": 0,
    }

    MAX_RETRIES = 5
    tentativa = 0

    async with httpx.AsyncClient(timeout=TIMEOUT) as client:
        while True:
            resp = await client.post(API_URL, headers=_anthropic_headers(cache), json=body)

            if resp.status_code == 429 and tentativa < MAX_RETRIES:
                tentativa += 1
                espera = int(resp.headers.get("retry-after", 60))
                print(f"   ⏳ Rate limit (429) — aguardando {espera}s (tentativa {tentativa}/{MAX_RETRIES})...")
                await asyncio.sleep(espera)
                continue

            resp.raise_for_status()
            data = resp.json()

            for k in uso_total:
                uso_total[k] += data.get("usage", {}).get(k, 0)

            if data.get("stop_reason") != "tool_use":
                return _juntar_texto(data.get("content", [])), uso_total

            # Loop tool_use — idêntico ao original
            tool_results = []
            for bloco in data["content"]:
                if bloco.get("type") == "tool_use":
                    tool_results.append({
                        "type":        "tool_result",
                        "tool_use_id": bloco["id"],
                        "content":     json.dumps(bloco.get("input", {})),
                    })

            body["messages"] = [
                *body["messages"],
                {"role": "assistant", "content": data["content"]},
                {"role": "user",      "content": tool_results},
            ]


# ══════════════════════════════════════════════════════════════════════════════
# OCR E COMPRESSÃO DE PDF  (idêntico ao pipeline original)
# ══════════════════════════════════════════════════════════════════════════════

def _comprimir_pdf(src: Path, dst: Path) -> Path:
    """Ghostscript primeiro; fallback para pymupdf — idêntico ao original."""
    try:
        subprocess.run(
            ["gs", "-sDEVICE=pdfwrite", "-dCompatibilityLevel=1.4",
             "-dPDFSETTINGS=/screen", "-dNOPAUSE", "-dQUIET", "-dBATCH",
             f"-sOutputFile={dst}", str(src)],
            check=True, capture_output=True, timeout=60,
        )
        return dst
    except (subprocess.CalledProcessError, FileNotFoundError):
        pass
    try:
        import fitz
        doc = fitz.open(str(src))
        doc_novo = fitz.open()
        for page in doc:
            mat = fitz.Matrix(100 / 72, 100 / 72)
            pix = page.get_pixmap(matrix=mat, colorspace=fitz.csGRAY)
            img_page = fitz.open()
            img_page.new_page(width=pix.width, height=pix.height)
            img_page[0].insert_image(img_page[0].rect, pixmap=pix)
            doc_novo.insert_pdf(img_page)
        doc_novo.save(str(dst))
        doc.close()
        doc_novo.close()
        return dst
    except Exception as e:
        print(f"   ⚠️  Erro ao comprimir {src.name}: {e}")
        return src


def _ocr_pdf(src: Path) -> bool:
    """
    Extrai texto de PDF via pymupdf + pytesseract — idêntico ao original.
    PDFs com fonte corrompida (ex: CISSPoder) são convertidos para imagem e OCR aplicado.
    Salva resultado como .txt na mesma pasta.
    """
    try:
        import fitz
        import pytesseract
        from PIL import Image
        import io

        doc = fitz.open(str(src))
        textos = []

        for page in doc:
            texto = page.get_text("text").strip()
            chars_validos = sum(1 for c in texto if c.isalnum() or c in " .,;:-R$/%\n")
            proporcao_valida = chars_validos / max(len(texto), 1)

            if len(texto) < 50 or proporcao_valida < 0.7:
                mat = fitz.Matrix(200 / 72, 200 / 72)
                pix = page.get_pixmap(matrix=mat)
                img_bytes = pix.tobytes("png")
                img = Image.open(io.BytesIO(img_bytes))
                texto = pytesseract.image_to_string(img, lang="por+eng", config="--psm 6")
                texto = texto.strip()

            if texto:
                textos.append(f"--- Página {page.number + 1} ---\n{texto}")

        doc.close()

        if textos:
            txt_path = src.with_suffix(".txt")
            txt_path.write_text("\n\n".join(textos), encoding="utf-8")
            return True
        return False

    except ImportError as e:
        print(f"   ⚠️  Dependência OCR faltando: {e}")
        return False
    except Exception as e:
        print(f"   ⚠️  OCR falhou em {src.name}: {e}")
        return False


def _consolidar_agendas_cerc(docs_dir: Path) -> None:
    """
    Funde múltiplos XLSX de agenda CERC em um único .txt — idêntico ao original.
    Arquivos nomeados agenda_*.xlsx.
    """
    try:
        import openpyxl
    except ImportError:
        print("   ⚠️  openpyxl não instalado")
        return

    agendas = sorted(docs_dir.glob("agenda_*.xlsx"))
    if not agendas:
        return

    print(f"   📊 {len(agendas)} agenda(s) CERC encontrada(s) — consolidando...")
    dados_cerc = []
    for xlsx in agendas:
        cnpj_filial = xlsx.stem.replace("agenda_", "")
        try:
            wb = openpyxl.load_workbook(xlsx, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                linhas = []
                for row in ws.iter_rows(values_only=True):
                    linha = "\t".join(str(c) if c is not None else "" for c in row)
                    if linha.strip():
                        linhas.append(linha)
                if linhas:
                    dados_cerc.append(f"=== CNPJ: {cnpj_filial} | Aba: {sheet} ===")
                    dados_cerc.extend(linhas[:20])
        except Exception as e:
            print(f"   ⚠️  {xlsx.name}: {e}")

    cerc_txt = docs_dir / "CERC_AGENDAS_CONSOLIDADO.txt"
    cerc_txt.write_text("\n".join(dados_cerc), encoding="utf-8")
    print(f"   ✅ CERC consolidado → {cerc_txt.name}")


def _preparar_documentos(docs_dir: Path) -> Path:
    """
    Comprime PDFs grandes, aplica OCR e consolida agendas CERC.
    Retorna path da pasta com documentos prontos.
    Espelha exatamente o pré-processamento do pipeline original.
    """
    comp_dir = docs_dir.parent / "documentos_comprimidos"
    comp_dir.mkdir(exist_ok=True)

    # Consolidar agendas CERC
    _consolidar_agendas_cerc(docs_dir)

    # Comprimir PDFs grandes
    print("   🗜️  Verificando PDFs...")
    for pdf in sorted(docs_dir.glob("*.pdf")):
        kb = pdf.stat().st_size / 1024
        dst = comp_dir / pdf.name
        if kb > LIMITE_KB_PDF:
            print(f"   🔧 {pdf.name} ({kb:.0f} KB) → comprimindo...")
            _comprimir_pdf(pdf, dst)
            if dst.exists():
                kb_novo = dst.stat().st_size / 1024
                print(f"      ✅ {kb_novo:.0f} KB ({(1 - kb_novo/kb)*100:.0f}% menor)")
            else:
                shutil.copy2(pdf, dst)
        else:
            shutil.copy2(pdf, dst)

    # Copiar outros arquivos (não-PDF, não-agenda)
    for f in docs_dir.iterdir():
        if f.suffix.lower() != ".pdf" and not f.name.startswith("agenda_") and f.is_file():
            shutil.copy2(f, comp_dir / f.name)

    # Aplicar OCR em PDFs
    print("   🔍 Aplicando OCR...")
    for pdf in sorted(comp_dir.glob("*.pdf")):
        txt = pdf.with_suffix(".txt")
        if txt.exists():
            print(f"   ⏭️  {pdf.name} — OCR já feito")
            continue
        print(f"   🔍 {pdf.name}...", end=" ")
        ok = _ocr_pdf(pdf)
        if ok:
            print(f"✅ → {pdf.stem}.txt")
        else:
            print("⚠️  falhou — enviará PDF original")

    return comp_dir


# ══════════════════════════════════════════════════════════════════════════════
# CARREGAMENTO DE DOCUMENTOS  (idêntico ao original)
# ══════════════════════════════════════════════════════════════════════════════

EXTENSOES_SUPORTADAS = {".pdf", ".png", ".jpg", ".jpeg", ".xlsx", ".txt"}
MIME = {
    ".pdf":  "application/pdf",
    ".png":  "image/png",
    ".jpg":  "image/jpeg",
    ".jpeg": "image/jpeg",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".txt":  "text/plain",
}
PRIORIDADE = [
    ["scr", "bacen"],
    ["serasa"],
    ["quod"],
    ["nuclea", "cip"],
    ["cerc", "agenda"],
    ["painel", "gestao", "gerencial", "consolidado"],
    ["balanco", "balanço", "balancete"],
    ["dre", "resultado"],
    ["faturamento", "fat"],
]


def _prioridade_doc(nome: str) -> int:
    n = nome.lower()
    for i, palavras in enumerate(PRIORIDADE):
        if any(p in n for p in palavras):
            return i
    return len(PRIORIDADE)


def _carregar_docs(docs_dir: Path) -> list[dict]:
    """
    Carrega arquivos da pasta e converte para content blocks.
    Lógica idêntica ao pipeline original (PDF/imagem/XLSX/OCR txt).
    """
    try:
        import openpyxl
        _tem_openpyxl = True
    except ImportError:
        _tem_openpyxl = False

    arquivos = sorted(
        [f for f in docs_dir.iterdir() if f.is_file() and f.suffix.lower() in EXTENSOES_SUPORTADAS],
        key=lambda f: (_prioridade_doc(f.name), f.name),
    )

    blocks = []
    for path in arquivos:
        ext  = path.suffix.lower()
        mime = MIME.get(ext, "")
        kb   = path.stat().st_size / 1024

        if kb > 1000:
            print(f"   ⚠️  {path.name} grande ({kb:.0f} KB) — pode estourar contexto")

        # PDF com OCR disponível → pular PDF, usar .txt
        if ext == ".pdf" and (path.parent / (path.stem + ".txt")).exists():
            print(f"   ⏭️  {path.name} — usando versão OCR .txt")
            continue

        if mime == "application/pdf":
            data = base64.standard_b64encode(path.read_bytes()).decode()
            blocks.append({
                "type": "document",
                "source": {"type": "base64", "media_type": "application/pdf", "data": data},
                "title": path.name,
            })
            print(f"   📄 {path.name} ({kb:.0f} KB) [PDF]")

        elif mime.startswith("image/"):
            data = base64.standard_b64encode(path.read_bytes()).decode()
            blocks.append({
                "type": "image",
                "source": {"type": "base64", "media_type": mime, "data": data},
            })
            print(f"   🖼️  {path.name} ({kb:.0f} KB) [imagem]")

        elif ext == ".xlsx":
            if not _tem_openpyxl:
                print(f"   ⚠️  {path.name} ignorado — instale openpyxl")
                continue
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, data_only=True)
                linhas = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    linhas.append(f"=== Aba: {sheet} ===")
                    for row in ws.iter_rows(values_only=True):
                        linha = "\t".join(str(c) if c is not None else "" for c in row)
                        if linha.strip():
                            linhas.append(linha)
                texto_xlsx = "\n".join(linhas)
                blocks.append({
                    "type": "document",
                    "source": {"type": "text", "media_type": "text/plain", "data": texto_xlsx},
                    "title": path.name,
                })
                print(f"   📊 {path.name} ({kb:.0f} KB) [XLSX→texto, {len(linhas)} linhas]")
            except Exception as e:
                print(f"   ⚠️  {path.name} ignorado — {e}")

        elif ext == ".txt":
            texto = path.read_text(encoding="utf-8", errors="replace")
            blocks.append({
                "type": "document",
                "source": {"type": "text", "media_type": "text/plain", "data": texto},
                "title": path.stem + ".pdf (OCR)",
            })
            print(f"   📝 {path.name} ({kb:.0f} KB) [OCR texto]")

        else:
            print(f"   ⏭️  {path.name} ignorado — formato não suportado")

    return blocks


# ══════════════════════════════════════════════════════════════════════════════
# FUSÃO DE P2 PARCIAIS  (idêntico ao original)
# ══════════════════════════════════════════════════════════════════════════════

def _fundir_p2(parciais: list[dict]) -> dict:
    if len(parciais) == 1:
        return parciais[0]
    base = parciais[0].copy()
    LISTAS = ["dre", "balanco", "faturamento_mensal",
               "inconsistencias_p2", "campos_ausentes_p2", "documentos_recebidos"]
    for p in parciais[1:]:
        for k, v in p.items():
            if k in LISTAS:
                existente = base.get(k) or []
                base[k] = existente + [x for x in (v or []) if x not in existente]
            elif k == "bureaux":
                bur = base.get("bureaux") or {}
                for bureau, dados in (v or {}).items():
                    if dados and not bur.get(bureau):
                        bur[bureau] = dados
                base["bureaux"] = bur
            elif k == "fontes_input":
                fi = base.get("fontes_input") or {}
                base["fontes_input"] = {
                    key: fi.get(key) or (v or {}).get(key)
                    for key in set(fi) | set(v or {})
                }
            elif base.get(k) is None and v is not None:
                base[k] = v
    return base


# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

async def _rodar_pipeline(analysis_id: str) -> None:
    """
    Executa o pipeline completo E1→E2→E3→E4→montador para um analysis_id.
    Salva resultados intermediários no Supabase a cada etapa.
    """
    print(f"\n{'='*60}")
    print(f"🚀 Iniciando pipeline — analysis_id: {analysis_id}")
    print(f"{'='*60}\n")

    try:
        # ── Setup: buscar registro e arquivos ─────────────────────────────
        rows = await _sb_get("analyses", f"id=eq.{analysis_id}&select=*")
        if not rows:
            raise RuntimeError(f"analysis_id {analysis_id} não encontrado")
        analise = rows[0]

        cnpj    = analise["cnpj"]
        produto = analise.get("produto")
        await _sb_patch("analyses", analysis_id, {"status": "running"})

        # Buscar documentos uploaded
        doc_rows = await _sb_get(
            "uploaded_documents",
            f"analysis_id=eq.{analysis_id}&select=*",
        )

        # Buscar prompts do Supabase
        p_e1_rows = await _sb_get("prompts", "name=eq.PROMPT_00_PESQUISA_v5&select=content")
        p_e2_rows = await _sb_get("prompts", "name=eq.PROMPT_01_EXTRACAO_v7&select=content")
        p_e4_rows = await _sb_get("prompts", "name=eq.PROMPT_04_MEMORANDO_v1&select=content")

        if not p_e1_rows or not p_e2_rows or not p_e4_rows:
            raise RuntimeError("Um ou mais prompts não encontrados na tabela 'prompts'")

        prompt_e1 = p_e1_rows[0]["content"]
        prompt_e2 = p_e2_rows[0]["content"]
        prompt_e4 = p_e4_rows[0]["content"]

        # Criar pasta temporária para os documentos
        with tempfile.TemporaryDirectory() as tmpdir:
            docs_orig = Path(tmpdir) / "documentos"
            docs_orig.mkdir()

            # Baixar arquivos do Supabase Storage
            print(f"⬇️  Baixando {len(doc_rows)} arquivo(s)...")
            for doc in doc_rows:
                file_bytes = await _sb_download(doc["storage_path"])
                dest = docs_orig / doc["file_name"]
                dest.write_bytes(file_bytes)
                print(f"   ✅ {doc['file_name']} ({len(file_bytes)//1024:.0f} KB)")

            # ── Pré-processar: compressão, OCR, CERC ─────────────────────
            print("\n🗜️  Pré-processando documentos...")
            docs_dir = _preparar_documentos(docs_orig)

            # ── E1: Pesquisa pública ──────────────────────────────────────
            print("\n🌐 E1 — Pesquisa pública...")
            t0 = time.monotonic()

            mensagem_e1 = [{
                "role": "user",
                "content": [{
                    "type": "text",
                    "text": f"CNPJ: {cnpj}\n\n{prompt_e1}",
                    "cache_control": {"type": "ephemeral"},
                }],
            }]

            texto_e1, uso_e1 = await _chamar_api(
                mensagens=mensagem_e1,
                max_tokens=MAX_TOKENS_E1,
                tools=[{"type": "web_search_20250305", "name": "web_search"}],
                cache=True,
                modelo=MODELO_E1,
            )

            p1 = _extrair_json(texto_e1)
            empresa = p1.get("company_name", cnpj)
            print(f"✅ E1 concluído em {time.monotonic()-t0:.1f}s | empresa: {empresa}")
            await _sb_patch("analyses", analysis_id, {"p1": p1})

            # ── E2: Extração de documentos ────────────────────────────────
            print("\n📄 E2 — Extração de documentos...")
            t0 = time.monotonic()

            docs_blocks = _carregar_docs(docs_dir)
            if not docs_blocks:
                raise RuntimeError(f"Nenhum documento encontrado em {docs_dir}")

            print(f"   Total: {len(docs_blocks)} documento(s)")

            instrucao_base = (
                f"Empresa: {empresa}\nCNPJ: {cnpj}\n"
                f"Produto prioritário: {produto or 'null'}\n"
            )

            p2_parciais = []
            uso_e2_total = {
                "input_tokens": 0, "output_tokens": 0,
                "cache_creation_input_tokens": 0, "cache_read_input_tokens": 0,
            }
            total_lotes = (len(docs_blocks) + CHUNK_DOCS - 1) // CHUNK_DOCS

            for i in range(0, len(docs_blocks), CHUNK_DOCS):
                lote = docs_blocks[i : i + CHUNK_DOCS]
                n = i // CHUNK_DOCS + 1
                print(f"   Lote {n}/{total_lotes}...")

                instrucao_lote = (
                    f"{instrucao_base}"
                    f"ATENÇÃO: lote {n} de {total_lotes}. "
                    f"Extraia apenas o que estiver nestes documentos. "
                    f"Campos ausentes neste lote: null.\n\n"
                )
                content = [
                    *lote,
                    {
                        "type": "text",
                        "text": instrucao_lote + prompt_e2,
                        "cache_control": {"type": "ephemeral"},
                    },
                ]

                texto_e2, uso_lote = await _chamar_api(
                    mensagens=[{"role": "user", "content": content}],
                    max_tokens=MAX_TOKENS_E2,
                    cache=True,
                    modelo=MODELO_E2,
                )
                p2_parciais.append(_extrair_json(texto_e2))
                for k in uso_e2_total:
                    uso_e2_total[k] += uso_lote.get(k, 0)
                print(f"   ✅ Lote {n} OK — out: {uso_lote['output_tokens']:,} tokens")

                if i + CHUNK_DOCS < len(docs_blocks):
                    print(f"   💤 Aguardando {SLEEP_ENTRE_LOTES}s (rate limit)...")
                    await asyncio.sleep(SLEEP_ENTRE_LOTES)

        # Fundir p2 parciais
        p2 = _fundir_p2(p2_parciais)
        if produto and not p2.get("produto_prioritario"):
            p2["produto_prioritario"] = produto

        print(f"✅ E2 concluído em {time.monotonic()-t0:.1f}s")

        # Validar p2 e logar alertas (não bloqueia)
        alertas = validar_p2(p2)
        for a in alertas:
            icone = "🔴" if a["nivel"] == "critico" else "🟡"
            print(f"   {icone} [{a['nivel'].upper()}] {a['campo']}: {a['impacto']}")

        await _sb_patch("analyses", analysis_id, {"p2": p2})

        # ── E3: Calculadora (Python puro — usa calculadora.py sem alteração) ──
        print("\n⚙️  E3 — Calculadora...")
        t0 = time.monotonic()
        p3 = calcular(p2)
        print(f"✅ E3 concluído em {time.monotonic()-t0:.2f}s — {len(p3)} blocos")
        await _sb_patch("analyses", analysis_id, {"p3": p3})

        # ── E4: Memorando ─────────────────────────────────────────────────
        print("\n📝 E4 — Memorando...")
        t0 = time.monotonic()

        p2_resumido = {k: v for k, v in p2.items() if k not in ("balanco", "dre", "faturamento_mensal")}

        mensagem_texto = (
            f"## dict_p1 (pesquisa pública)\n"
            f"```json\n{json.dumps(p1, ensure_ascii=False, indent=2)}\n```\n\n"
            f"## dict_p2 (bureaux e metadados — sem raw financeiro)\n"
            f"```json\n{json.dumps(p2_resumido, ensure_ascii=False, indent=2)}\n```\n\n"
            f"## dict_p3 (indicadores calculados)\n"
            f"```json\n{json.dumps(p3, ensure_ascii=False, indent=2)}\n```\n\n"
            f"---\n\n{prompt_e4}"
        )

        mensagem_e4 = [{
            "role": "user",
            "content": [{
                "type": "text",
                "text": mensagem_texto,
                "cache_control": {"type": "ephemeral"},
            }],
        }]

        texto_e4, uso_e4 = await _chamar_api(
            mensagens=mensagem_e4,
            max_tokens=MAX_TOKENS_E4,
            cache=True,
            modelo=MODELO_E4,
        )

        p4 = _extrair_json(texto_e4)
        print(f"✅ E4 concluído em {time.monotonic()-t0:.1f}s")
        await _sb_patch("analyses", analysis_id, {"p4": p4})

        # ── E5: Montador + validação + envio ──────────────────────────────
        print("\n🔧 E5 — Montador e envio...")
        payload = montar_payload(p1, p2, p3, p4)

        todos  = validar_payload(payload)
        erros  = [e for e in todos if not e.startswith("⚠️")]
        avisos = [e for e in todos if e.startswith("⚠️")]

        info = inspecionar_payload(payload)
        print(f"   Tamanho:  {info['tamanho_payload_kb']} KB")
        print(f"   Slides:   {info['slides_presentes']} presentes, {info['slides_nulos']} nulos")
        for a in avisos:
            print(f"   {a}")

        if erros:
            print(f"⚠️  Avisos de validação (não bloqueante): {'; '.join(erros)}")

        if LOVABLE_API_KEY:
            resultado = await enviar_payload(payload, api_key=LOVABLE_API_KEY)
            if resultado.sucesso:
                print(f"✅ Payload enviado — HTTP {resultado.status_code} em {resultado.duracao_s}s")
            else:
                print(f"⚠️  Falha no envio: {resultado.erro} (análise salva no Supabase)")
        else:
            print("ℹ️  LOVABLE_API_KEY não definida — payload não enviado ao endpoint externo")

        # Marcar como concluído
        await _sb_patch("analyses", analysis_id, {"status": "done"})
        print(f"\n✅ Pipeline concluído — analysis_id: {analysis_id}")

    except Exception as exc:
        print(f"\n❌ Erro no pipeline: {exc}")
        try:
            await _sb_patch("analyses", analysis_id, {
                "status": "error",
                "error_message": str(exc)[:2000],
            })
        except Exception as patch_exc:
            print(f"   ⚠️  Não foi possível salvar erro no Supabase: {patch_exc}")
        raise


# ══════════════════════════════════════════════════════════════════════════════
# ENDPOINTS FASTAPI
# ══════════════════════════════════════════════════════════════════════════════

class RunPipelineRequest(BaseModel):
    analysis_id: str


@app.post("/run-pipeline")
async def run_pipeline(
    body: RunPipelineRequest,
    background_tasks: BackgroundTasks,
    x_pipeline_secret: str | None = Header(default=None),
):
    """
    Dispara o pipeline em background e retorna 202 imediatamente.
    A Supabase Edge Function chama este endpoint passando o header X-Pipeline-Secret.
    """
    if PIPELINE_SECRET and x_pipeline_secret != PIPELINE_SECRET:
        raise HTTPException(status_code=401, detail="Unauthorized")

    background_tasks.add_task(_rodar_pipeline, body.analysis_id)
    return JSONResponse(status_code=202, content={"status": "accepted", "analysis_id": body.analysis_id})


@app.get("/health")
async def health():
    return {"status": "ok"}
